"""
OneDrive integration for apartment listing storage.

Reads and writes an Excel file on OneDrive via the Microsoft Graph API,
using the same MSAL refresh-token pattern as personal_log_etl_cloud.py.

Key difference from the read-only ETL: this module needs Files.ReadWrite
scope so it can upload the updated Excel file back to OneDrive.

IMPORTANT — first-time setup:
  Run setup_onedrive_auth.py with SCOPES updated to include
  "https://graph.microsoft.com/Files.ReadWrite" to generate a new
  refresh token with write permissions. Store it as a separate GitHub
  secret (APARTMENT_ONEDRIVE_REFRESH_TOKEN) if you want to keep the
  personal log token unchanged.

Environment variables required:
  AZURE_CLIENT_ID                   — same as personal log
  AZURE_TENANT_ID                   — same as personal log
  APARTMENT_ONEDRIVE_REFRESH_TOKEN  — refresh token with Files.ReadWrite
  APARTMENT_ONEDRIVE_FILE_PATH      — path in OneDrive, e.g.
                                      "Documents/Apartment Listings.xlsx"
                                      (overrides config.yaml value)
"""

import io
import os
import tempfile
from typing import Optional

import msal
import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

from apartment_hunter.models import Listing, EXCEL_COLUMNS

SCOPES = ["https://graph.microsoft.com/Files.ReadWrite"]
GRAPH_BASE = "https://graph.microsoft.com/v1.0/me/drive/root:/{path}:{action}"


def download_listings(config: dict) -> pd.DataFrame:
    """
    Download the apartment listings Excel from OneDrive and return a DataFrame.
    Returns an empty DataFrame (with correct columns) if the file doesn't exist yet.
    """
    token = _get_token()
    file_path = _resolve_file_path(config)

    url = GRAPH_BASE.format(path=file_path, action="/content")
    resp = requests.get(url, headers={"Authorization": f"Bearer {token}"}, timeout=30)

    if resp.status_code == 404:
        print(f"  [OneDrive] '{file_path}' not found — will create on first upload")
        return _empty_df()

    resp.raise_for_status()

    sheet = config.get("onedrive", {}).get("sheet_name", "Listings")
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(resp.content)
        tmp_path = tmp.name

    try:
        df = pd.read_excel(tmp_path, sheet_name=sheet, engine="openpyxl", dtype=str)
        # Ensure all expected columns exist (handles schema additions gracefully)
        for col in EXCEL_COLUMNS:
            if col not in df.columns:
                df[col] = None
        df = df[EXCEL_COLUMNS]
        print(f"  [OneDrive] loaded {len(df)} existing listings")
        return df
    except Exception as e:
        print(f"  [OneDrive] error reading Excel: {e}")
        return _empty_df()
    finally:
        os.remove(tmp_path)


def merge_listings(existing_df: pd.DataFrame, new_listings: list[Listing]) -> pd.DataFrame:
    """
    Append new Listing objects to the existing DataFrame.
    new_listings should already be de-duplicated against existing_df by the caller.
    """
    if not new_listings:
        return existing_df

    new_rows = pd.DataFrame([l.to_dict() for l in new_listings])
    new_rows = new_rows[EXCEL_COLUMNS]

    merged = pd.concat([existing_df, new_rows], ignore_index=True)
    print(f"  [OneDrive] merged: {len(existing_df)} existing + {len(new_listings)} new = {len(merged)} total")
    return merged


def upload_listings(config: dict, df: pd.DataFrame) -> None:
    """
    Write the DataFrame to an Excel file and upload it to OneDrive,
    overwriting the existing file.
    """
    token = _get_token()
    file_path = _resolve_file_path(config)
    sheet = config.get("onedrive", {}).get("sheet_name", "Listings")

    excel_bytes = _df_to_excel_bytes(df, sheet)

    # Graph API upload endpoint (handles both create and overwrite)
    url = GRAPH_BASE.format(path=file_path, action="/content")
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }
    resp = requests.put(url, headers=headers, data=excel_bytes, timeout=60)
    resp.raise_for_status()
    print(f"  [OneDrive] uploaded {len(df)} listings to '{file_path}'")


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _get_token() -> str:
    client_id    = os.environ["AZURE_CLIENT_ID"]
    tenant_id    = os.environ.get("AZURE_TENANT_ID", "common")
    refresh_token = os.environ["APARTMENT_ONEDRIVE_REFRESH_TOKEN"]

    app = msal.PublicClientApplication(
        client_id=client_id,
        authority=f"https://login.microsoftonline.com/{tenant_id}",
    )
    result = app.acquire_token_by_refresh_token(refresh_token, scopes=SCOPES)
    if "access_token" not in result:
        raise SystemExit(
            f"Could not acquire OneDrive token.\n"
            f"Error: {result.get('error')}\n"
            f"Details: {result.get('error_description')}\n\n"
            "Re-run setup_onedrive_auth.py with Files.ReadWrite scope and update "
            "the APARTMENT_ONEDRIVE_REFRESH_TOKEN secret."
        )
    return result["access_token"]


def _resolve_file_path(config: dict) -> str:
    """
    File path priority:
      1. APARTMENT_ONEDRIVE_FILE_PATH env var
      2. config.yaml onedrive.file_path
      3. Default
    """
    return (
        os.environ.get("APARTMENT_ONEDRIVE_FILE_PATH")
        or config.get("onedrive", {}).get("file_path", "Documents/Apartment Listings.xlsx")
    )


def _empty_df() -> pd.DataFrame:
    return pd.DataFrame(columns=EXCEL_COLUMNS)


def _df_to_excel_bytes(df: pd.DataFrame, sheet_name: str) -> bytes:
    """Convert a DataFrame to styled Excel bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2C3E50")
    priority_fill = PatternFill("solid", fgColor="FFF3CD")

    # Write header row
    for col_idx, col_name in enumerate(EXCEL_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Write data rows
    for row_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

        # Highlight priority rows
        is_priority_col = EXCEL_COLUMNS.index("is_priority")
        priority_val = str(row[is_priority_col]).lower() if row[is_priority_col] is not None else ""
        if priority_val in ("true", "1", "yes"):
            for col_idx in range(1, len(EXCEL_COLUMNS) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = priority_fill

    # Auto-fit column widths (approximate)
    col_widths = {
        "listing_id": 14, "title": 45, "source": 14, "price": 10,
        "neighborhood": 22, "address": 35, "floor": 8, "bedrooms": 10,
        "bathrooms": 10, "rent_stabilized": 16, "date_listed": 16,
        "nearest_subway": 45, "date_found": 16, "url": 50, "is_priority": 12, "reviewed": 10,
    }
    for col_idx, col_name in enumerate(EXCEL_COLUMNS, start=1):
        ws.column_dimensions[_col_letter(col_idx)].width = col_widths.get(col_name, 15)

    # Freeze header row
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _col_letter(n: int) -> str:
    """Convert 1-based column index to Excel letter (A, B, ..., Z, AA, ...)."""
    result = ""
    while n:
        n, rem = divmod(n - 1, 26)
        result = chr(65 + rem) + result
    return result
